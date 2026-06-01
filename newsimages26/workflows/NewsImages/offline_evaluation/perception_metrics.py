# -*- coding: utf-8 -*-
r"""
Full perceptual metric script for AI-generated images.

Generated image folder:
    D:\Users\Desktop\newsimages_survey_26_v1.1\testresult20260527-Flux2-Klein-4b-base

Real reference image folder:
    D:\Users\Desktop\newsimages_survey_26_v1.1\news_images_survey

Output:
    no_reference_perception_metrics_full.xlsx

This version avoids torchmetrics / torch-fidelity because old Anaconda environments
may fail with:
    cannot import name 'OrderedDict' from 'typing'

Metrics:
    1. pyiqa no-reference metrics, if available.
    2. Optional prompt-based metrics, if packages are available.
    3. Auxiliary image diagnostics.
    4. Custom FID/KID using torchvision Inception-v3 features.
"""

import os
import gc
import math
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image

warnings.filterwarnings("ignore")


# ============================================================
# User settings
# ============================================================

IMAGE_DIR = Path(
    r"D:\Users\Desktop\newsimages_survey_26_v1.1\testresult20260527-Flux2-Klein-4b-base"
)

REAL_IMAGE_DIR = Path(
    r"D:\Users\Desktop\newsimages_survey_26_v1.1\news_images_survey"
)

OUTPUT_XLSX = IMAGE_DIR / "no_reference_perception_metrics_full.xlsx"

IMAGE_EXTS = {".png"}
REAL_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".webp"}

RECURSIVE_REAL_SCAN = True

USE_CUDA_IF_AVAILABLE = True

GLOBAL_PROMPT = (
    "restore and recolor this old photograph, add camera imperfections and washed out colors, "
    "dull skin and not glossy"
)

# Optional prompt metrics.
# If your environment is fragile, you can set these to False.
ENABLE_CLIPSCORE = True
ENABLE_IMAGEREWARD = True
ENABLE_PICKSCORE = True
ENABLE_HPSV2 = True

# Custom FID/KID.
ENABLE_FID_KID = True
FID_KID_BATCH_SIZE = 8
KID_SUBSETS = 50
KID_SUBSET_SIZE = 50
RANDOM_SEED = 123

# pyiqa no-reference metrics.
PYIQA_NR_METRICS = [
    "niqe",
    "brisque",
    "pi",
    "nrqm",
    "musiq",
    "maniqa",
    "clipiqa",
    "clipiqa+",
    "dbcnn",
    "nima",
    "hyperiqa",
    "paq2piq",
    "topiq_nr",
    "liqe",
    "cnniqa",
    "tres",
]

PYIQA_COLUMNS = ["pyiqa_{}".format(x) for x in PYIQA_NR_METRICS]

PROMPT_COLUMNS = [
    "clipscore",
    "imagereward",
    "pickscore",
    "hpsv2",
]

PRIMARY_METRIC_COLUMNS = PYIQA_COLUMNS + PROMPT_COLUMNS


# ============================================================
# Basic utilities
# ============================================================

def collect_images(folder, exts, recursive=False):
    if not folder.exists():
        raise FileNotFoundError("Folder does not exist: {}".format(folder))

    if recursive:
        iterator = folder.rglob("*")
    else:
        iterator = folder.iterdir()

    paths = []
    for p in iterator:
        if p.is_file() and p.suffix.lower() in exts:
            paths.append(p)

    paths = sorted(paths, key=lambda x: str(x).lower())
    return paths


def safe_float(x):
    try:
        if hasattr(x, "detach"):
            x = x.detach().cpu().numpy()

        if isinstance(x, list) or isinstance(x, tuple):
            if len(x) == 0:
                return np.nan
            x = x[0]

        if isinstance(x, np.ndarray):
            x = float(np.mean(x))

        return float(x)
    except Exception:
        return np.nan


def get_torch_device():
    try:
        import torch
        if USE_CUDA_IF_AVAILABLE and torch.cuda.is_available():
            return torch.device("cuda")
        return torch.device("cpu")
    except Exception:
        return None


def release_memory():
    gc.collect()
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
    except Exception:
        pass


def load_rgb_array(image_path):
    img = Image.open(str(image_path)).convert("RGB")
    return np.asarray(img).astype(np.float32)


def rgb_to_gray(rgb):
    r = rgb[:, :, 0]
    g = rgb[:, :, 1]
    b = rgb[:, :, 2]
    return (0.299 * r + 0.587 * g + 0.114 * b).astype(np.float32)


# ============================================================
# Auxiliary image diagnostics
# ============================================================

def compute_entropy(gray):
    hist, _ = np.histogram(gray, bins=256, range=(0, 255), density=True)
    hist = hist[hist > 0]
    if len(hist) == 0:
        return np.nan
    return float(-np.sum(hist * np.log2(hist)))


def compute_dynamic_range(gray):
    return float(np.percentile(gray, 99) - np.percentile(gray, 1))


def compute_laplacian_variance_numpy(gray):
    g = gray.astype(np.float32)

    if g.shape[0] < 3 or g.shape[1] < 3:
        return np.nan

    lap = (
        -4.0 * g[1:-1, 1:-1]
        + g[:-2, 1:-1]
        + g[2:, 1:-1]
        + g[1:-1, :-2]
        + g[1:-1, 2:]
    )

    return float(np.var(lap))


def compute_tenengrad_numpy(gray):
    g = gray.astype(np.float32)

    if g.shape[0] < 3 or g.shape[1] < 3:
        return np.nan

    gx = (
        -1 * g[:-2, :-2] + 1 * g[:-2, 2:]
        -2 * g[1:-1, :-2] + 2 * g[1:-1, 2:]
        -1 * g[2:, :-2] + 1 * g[2:, 2:]
    )

    gy = (
        -1 * g[:-2, :-2] - 2 * g[:-2, 1:-1] - 1 * g[:-2, 2:]
        +1 * g[2:, :-2] + 2 * g[2:, 1:-1] + 1 * g[2:, 2:]
    )

    return float(np.mean(gx * gx + gy * gy))


def compute_edge_density_numpy(gray):
    g = gray.astype(np.float32)

    gx = np.zeros_like(g)
    gy = np.zeros_like(g)

    gx[:, 1:-1] = g[:, 2:] - g[:, :-2]
    gy[1:-1, :] = g[2:, :] - g[:-2, :]

    mag = np.sqrt(gx * gx + gy * gy)
    threshold = np.percentile(mag, 90)

    if threshold <= 1e-8:
        return 0.0

    return float(np.mean(mag > threshold))


def compute_colorfulness(rgb):
    r = rgb[:, :, 0]
    g = rgb[:, :, 1]
    b = rgb[:, :, 2]

    rg = np.abs(r - g)
    yb = np.abs(0.5 * (r + g) - b)

    value = np.sqrt(np.std(rg) ** 2 + np.std(yb) ** 2) + \
        0.3 * np.sqrt(np.mean(rg) ** 2 + np.mean(yb) ** 2)

    return float(value)


def compute_saturation_mean_numpy(rgb):
    rgb01 = np.clip(rgb / 255.0, 0.0, 1.0)

    maxc = np.max(rgb01, axis=2)
    minc = np.min(rgb01, axis=2)

    sat = np.zeros_like(maxc)
    mask = maxc > 1e-8
    sat[mask] = (maxc[mask] - minc[mask]) / maxc[mask]

    return float(np.mean(sat))


def try_compute_laplacian_variance(gray):
    try:
        import cv2
        lap = cv2.Laplacian(gray.astype(np.float32), cv2.CV_32F)
        return float(lap.var())
    except Exception:
        return compute_laplacian_variance_numpy(gray)


def try_compute_tenengrad(gray):
    try:
        import cv2
        gx = cv2.Sobel(gray.astype(np.float32), cv2.CV_32F, 1, 0, ksize=3)
        gy = cv2.Sobel(gray.astype(np.float32), cv2.CV_32F, 0, 1, ksize=3)
        return float(np.mean(gx * gx + gy * gy))
    except Exception:
        return compute_tenengrad_numpy(gray)


def try_compute_edge_density(gray):
    try:
        import cv2
        gray_u8 = np.clip(gray, 0, 255).astype(np.uint8)
        edges = cv2.Canny(gray_u8, 100, 200)
        return float(np.mean(edges > 0))
    except Exception:
        return compute_edge_density_numpy(gray)


def try_compute_saturation_mean(rgb):
    try:
        import cv2
        hsv = cv2.cvtColor(rgb.astype(np.uint8), cv2.COLOR_RGB2HSV)
        return float(np.mean(hsv[:, :, 1].astype(np.float32)))
    except Exception:
        return compute_saturation_mean_numpy(rgb)


def compute_auxiliary_metrics(image_path):
    rgb = load_rgb_array(image_path)
    gray = rgb_to_gray(rgb)

    h, w = gray.shape

    metrics = {}

    metrics["width"] = int(w)
    metrics["height"] = int(h)
    metrics["num_pixels"] = int(w * h)

    metrics["brightness_mean"] = float(np.mean(gray))
    metrics["brightness_std"] = float(np.std(gray))
    metrics["brightness_median"] = float(np.median(gray))
    metrics["brightness_min"] = float(np.min(gray))
    metrics["brightness_max"] = float(np.max(gray))
    metrics["dynamic_range_p99_p1"] = compute_dynamic_range(gray)

    metrics["dark_pixel_ratio_lt_10"] = float(np.mean(gray < 10))
    metrics["dark_pixel_ratio_lt_20"] = float(np.mean(gray < 20))
    metrics["bright_pixel_ratio_gt_235"] = float(np.mean(gray > 235))
    metrics["bright_pixel_ratio_gt_245"] = float(np.mean(gray > 245))

    metrics["laplacian_variance_sharpness"] = try_compute_laplacian_variance(gray)
    metrics["tenengrad_sharpness"] = try_compute_tenengrad(gray)
    metrics["edge_density"] = try_compute_edge_density(gray)

    metrics["entropy"] = compute_entropy(gray)
    metrics["colorfulness"] = compute_colorfulness(rgb)
    metrics["saturation_mean"] = try_compute_saturation_mean(rgb)

    metrics["red_mean"] = float(np.mean(rgb[:, :, 0]))
    metrics["green_mean"] = float(np.mean(rgb[:, :, 1]))
    metrics["blue_mean"] = float(np.mean(rgb[:, :, 2]))

    metrics["red_std"] = float(np.std(rgb[:, :, 0]))
    metrics["green_std"] = float(np.std(rgb[:, :, 1]))
    metrics["blue_std"] = float(np.std(rgb[:, :, 2]))

    metrics["rg_mean_abs_diff"] = float(np.mean(np.abs(rgb[:, :, 0] - rgb[:, :, 1])))
    metrics["gb_mean_abs_diff"] = float(np.mean(np.abs(rgb[:, :, 1] - rgb[:, :, 2])))
    metrics["rb_mean_abs_diff"] = float(np.mean(np.abs(rgb[:, :, 0] - rgb[:, :, 2])))

    return metrics


# ============================================================
# pyiqa no-reference metrics
# ============================================================

def init_pyiqa_metrics(metric_names):
    metric_objects = {}
    status_rows = []

    try:
        import pyiqa
    except Exception as e:
        for name in metric_names:
            status_rows.append({
                "metric": "pyiqa_{}".format(name),
                "group": "pyiqa no-reference IQA",
                "status": "failed",
                "note": "pyiqa unavailable: {}".format(repr(e))
            })
        return metric_objects, status_rows

    device = get_torch_device()

    if device is None:
        for name in metric_names:
            status_rows.append({
                "metric": "pyiqa_{}".format(name),
                "group": "pyiqa no-reference IQA",
                "status": "failed",
                "note": "torch device unavailable"
            })
        return metric_objects, status_rows

    print("[info] pyiqa device: {}".format(device))

    available_models = None
    try:
        available_models = set(pyiqa.list_models())
    except Exception:
        available_models = None

    for name in metric_names:
        col = "pyiqa_{}".format(name)

        try:
            if available_models is not None and name not in available_models:
                status_rows.append({
                    "metric": col,
                    "group": "pyiqa no-reference IQA",
                    "status": "skipped",
                    "note": "not listed in installed pyiqa version"
                })
                print("[skip] {}: not listed in installed pyiqa version".format(name))
                continue

            metric = pyiqa.create_metric(name, device=device)
            metric_objects[name] = metric

            status_rows.append({
                "metric": col,
                "group": "pyiqa no-reference IQA",
                "status": "loaded",
                "note": ""
            })

            print("[ok] loaded pyiqa metric: {}".format(name))

        except Exception as e:
            status_rows.append({
                "metric": col,
                "group": "pyiqa no-reference IQA",
                "status": "failed",
                "note": repr(e)
            })
            print("[skip] {}: {}".format(name, repr(e)))

    return metric_objects, status_rows


def compute_pyiqa_metrics(image_path, metric_objects):
    results = {}

    for name in PYIQA_NR_METRICS:
        results["pyiqa_{}".format(name)] = np.nan

    for name, metric in metric_objects.items():
        col = "pyiqa_{}".format(name)

        try:
            score = metric(str(image_path))
            results[col] = safe_float(score)
        except Exception as e:
            print("[warning] failed {} on {}: {}".format(name, image_path.name, repr(e)))
            results[col] = np.nan

    return results


# ============================================================
# Prompt-based metrics
# ============================================================

class CLIPScoreScorer(object):
    def __init__(self):
        self.available = False
        self.device = get_torch_device()
        self.note = ""

        try:
            import torch
            from transformers import CLIPProcessor, CLIPModel

            self.torch = torch
            self.Image = Image

            model_name = "openai/clip-vit-large-patch14"
            print("[info] loading CLIPScore model: {}".format(model_name))

            self.processor = CLIPProcessor.from_pretrained(model_name)
            self.model = CLIPModel.from_pretrained(model_name)
            self.model.to(self.device)
            self.model.eval()

            self.available = True
            print("[ok] CLIPScore loaded")

        except Exception as e:
            self.note = repr(e)
            print("[skip] CLIPScore unavailable: {}".format(repr(e)))

    def score(self, image_path, prompt):
        if not self.available:
            return np.nan

        try:
            image = self.Image.open(str(image_path)).convert("RGB")

            inputs = self.processor(
                text=[str(prompt)],
                images=image,
                return_tensors="pt",
                padding=True,
                truncation=True
            )

            for k in inputs:
                inputs[k] = inputs[k].to(self.device)

            with self.torch.no_grad():
                outputs = self.model(**inputs)
                image_embeds = outputs.image_embeds
                text_embeds = outputs.text_embeds

                image_embeds = image_embeds / image_embeds.norm(dim=-1, keepdim=True)
                text_embeds = text_embeds / text_embeds.norm(dim=-1, keepdim=True)

                score = 100.0 * (image_embeds * text_embeds).sum(dim=-1).item()

            return float(score)

        except Exception as e:
            print("[warning] CLIPScore failed on {}: {}".format(image_path.name, repr(e)))
            return np.nan


class ImageRewardScorer(object):
    def __init__(self):
        self.available = False
        self.note = ""

        try:
            import ImageReward as RM
            print("[info] loading ImageReward model")
            self.model = RM.load("ImageReward-v1.0")
            self.available = True
            print("[ok] ImageReward loaded")
        except Exception as e:
            self.note = repr(e)
            print("[skip] ImageReward unavailable: {}".format(repr(e)))

    def score(self, image_path, prompt):
        if not self.available:
            return np.nan

        try:
            return safe_float(self.model.score(str(prompt), str(image_path)))
        except Exception as e:
            print("[warning] ImageReward failed on {}: {}".format(image_path.name, repr(e)))
            return np.nan


class PickScoreScorer(object):
    def __init__(self):
        self.available = False
        self.device = get_torch_device()
        self.note = ""

        try:
            import torch
            from transformers import AutoProcessor, AutoModel

            self.torch = torch
            self.Image = Image

            model_name = "yuvalkirstain/PickScore_v1"
            processor_name = "laion/CLIP-ViT-H-14-laion2B-s32B-b79K"

            print("[info] loading PickScore model: {}".format(model_name))

            self.processor = AutoProcessor.from_pretrained(processor_name)
            self.model = AutoModel.from_pretrained(model_name)
            self.model.to(self.device)
            self.model.eval()

            self.available = True
            print("[ok] PickScore loaded")

        except Exception as e:
            self.note = repr(e)
            print("[skip] PickScore unavailable: {}".format(repr(e)))

    def score(self, image_path, prompt):
        if not self.available:
            return np.nan

        try:
            image = self.Image.open(str(image_path)).convert("RGB")

            image_inputs = self.processor(images=image, return_tensors="pt")
            text_inputs = self.processor(
                text=[str(prompt)],
                padding=True,
                truncation=True,
                max_length=77,
                return_tensors="pt"
            )

            for k in image_inputs:
                image_inputs[k] = image_inputs[k].to(self.device)
            for k in text_inputs:
                text_inputs[k] = text_inputs[k].to(self.device)

            with self.torch.no_grad():
                image_embeds = self.model.get_image_features(**image_inputs)
                text_embeds = self.model.get_text_features(**text_inputs)

                image_embeds = image_embeds / image_embeds.norm(dim=-1, keepdim=True)
                text_embeds = text_embeds / text_embeds.norm(dim=-1, keepdim=True)

                score = (image_embeds * text_embeds).sum(dim=-1).item()

            return float(score)

        except Exception as e:
            print("[warning] PickScore failed on {}: {}".format(image_path.name, repr(e)))
            return np.nan


class HPSv2Scorer(object):
    def __init__(self):
        self.available = False
        self.note = ""

        try:
            import hpsv2
            self.hpsv2 = hpsv2
            self.available = True
            print("[ok] HPSv2 package loaded")
        except Exception as e:
            self.note = repr(e)
            print("[skip] HPSv2 unavailable: {}".format(repr(e)))

    def score(self, image_path, prompt):
        if not self.available:
            return np.nan

        try:
            score = self.hpsv2.score(
                [str(image_path)],
                [str(prompt)],
                hps_version="v2.1"
            )
            return safe_float(score)
        except Exception as e:
            print("[warning] HPSv2 failed on {}: {}".format(image_path.name, repr(e)))
            return np.nan


def init_prompt_scorers():
    scorers = {}
    status_rows = []

    print("[info] global prompt:")
    print("       {}".format(GLOBAL_PROMPT))

    if ENABLE_CLIPSCORE:
        s = CLIPScoreScorer()
        if s.available:
            scorers["clipscore"] = s
            status_rows.append({"metric": "clipscore", "group": "prompt-based", "status": "loaded", "note": ""})
        else:
            status_rows.append({"metric": "clipscore", "group": "prompt-based", "status": "failed", "note": s.note})
        release_memory()

    if ENABLE_IMAGEREWARD:
        s = ImageRewardScorer()
        if s.available:
            scorers["imagereward"] = s
            status_rows.append({"metric": "imagereward", "group": "prompt-based", "status": "loaded", "note": ""})
        else:
            status_rows.append({"metric": "imagereward", "group": "prompt-based", "status": "failed", "note": s.note})
        release_memory()

    if ENABLE_PICKSCORE:
        s = PickScoreScorer()
        if s.available:
            scorers["pickscore"] = s
            status_rows.append({"metric": "pickscore", "group": "prompt-based", "status": "loaded", "note": ""})
        else:
            status_rows.append({"metric": "pickscore", "group": "prompt-based", "status": "failed", "note": s.note})
        release_memory()

    if ENABLE_HPSV2:
        s = HPSv2Scorer()
        if s.available:
            scorers["hpsv2"] = s
            status_rows.append({"metric": "hpsv2", "group": "prompt-based", "status": "loaded", "note": ""})
        else:
            status_rows.append({"metric": "hpsv2", "group": "prompt-based", "status": "failed", "note": s.note})
        release_memory()

    return scorers, status_rows


def compute_prompt_metrics(image_path, scorers):
    results = {}
    for col in PROMPT_COLUMNS:
        results[col] = np.nan

    for name, scorer in scorers.items():
        try:
            results[name] = scorer.score(image_path, GLOBAL_PROMPT)
        except Exception as e:
            print("[warning] {} failed on {}: {}".format(name, image_path.name, repr(e)))
            results[name] = np.nan

    return results


# ============================================================
# Custom FID / KID without torchmetrics
# ============================================================

class InceptionFeatureExtractor(object):
    def __init__(self):
        self.available = False
        self.note = ""
        self.device = get_torch_device()

        try:
            import torch
            import torchvision.models as models

            self.torch = torch

            print("[info] loading torchvision Inception-v3 for custom FID/KID")

            model = None

            # New torchvision API
            try:
                weights = models.Inception_V3_Weights.DEFAULT
                model = models.inception_v3(weights=weights, transform_input=False)
            except Exception:
                # Old torchvision API
                model = models.inception_v3(pretrained=True, transform_input=False)

            # Replace classifier with identity to get 2048-dim features.
            class Identity(torch.nn.Module):
                def forward(self, x):
                    return x

            model.fc = Identity()
            model.eval()
            model.to(self.device)

            self.model = model
            self.available = True

            print("[ok] Inception-v3 loaded for custom FID/KID")

        except Exception as e:
            self.note = repr(e)
            self.available = False
            print("[warning] custom FID/KID Inception loading failed: {}".format(repr(e)))

    def preprocess_batch(self, image_paths):
        torch = self.torch
        tensors = []

        mean = np.array([0.485, 0.456, 0.406], dtype=np.float32).reshape(1, 1, 3)
        std = np.array([0.229, 0.224, 0.225], dtype=np.float32).reshape(1, 1, 3)

        for p in image_paths:
            img = Image.open(str(p)).convert("RGB")
            img = img.resize((299, 299), Image.BICUBIC)

            arr = np.asarray(img).astype(np.float32) / 255.0
            arr = (arr - mean) / std

            arr = np.transpose(arr, (2, 0, 1))
            tensor = torch.from_numpy(arr).float()
            tensors.append(tensor)

        batch = torch.stack(tensors, dim=0)
        return batch

    def extract(self, image_paths, batch_size=8):
        if not self.available:
            raise RuntimeError("Inception extractor unavailable: {}".format(self.note))

        torch = self.torch
        all_features = []

        n = len(image_paths)

        for start in range(0, n, batch_size):
            end = min(start + batch_size, n)
            batch_paths = image_paths[start:end]

            print("[fid/kid] extracting features {}/{}".format(end, n))

            batch = self.preprocess_batch(batch_paths)
            batch = batch.to(self.device)

            with torch.no_grad():
                feat = self.model(batch)

                if isinstance(feat, tuple):
                    feat = feat[0]

                if len(feat.shape) > 2:
                    feat = torch.flatten(feat, 1)

            feat_np = feat.detach().cpu().numpy().astype(np.float64)
            all_features.append(feat_np)

            del batch, feat
            release_memory()

        features = np.concatenate(all_features, axis=0)
        return features


def calculate_fid_from_features(fake_features, real_features):
    from scipy import linalg

    mu_fake = np.mean(fake_features, axis=0)
    mu_real = np.mean(real_features, axis=0)

    sigma_fake = np.cov(fake_features, rowvar=False)
    sigma_real = np.cov(real_features, rowvar=False)

    diff = mu_fake - mu_real

    covmean, _ = linalg.sqrtm(sigma_fake.dot(sigma_real), disp=False)

    if not np.isfinite(covmean).all():
        eps = 1e-6
        offset = np.eye(sigma_fake.shape[0]) * eps
        covmean = linalg.sqrtm((sigma_fake + offset).dot(sigma_real + offset))

    if np.iscomplexobj(covmean):
        covmean = covmean.real

    fid = diff.dot(diff) + np.trace(sigma_fake + sigma_real - 2.0 * covmean)
    return float(fid)


def polynomial_kernel(x, y):
    dim = x.shape[1]
    return (np.dot(x, y.T) / float(dim) + 1.0) ** 3


def mmd2_unbiased(x, y):
    m = x.shape[0]
    n = y.shape[0]

    k_xx = polynomial_kernel(x, x)
    k_yy = polynomial_kernel(y, y)
    k_xy = polynomial_kernel(x, y)

    sum_xx = (np.sum(k_xx) - np.trace(k_xx)) / float(m * (m - 1))
    sum_yy = (np.sum(k_yy) - np.trace(k_yy)) / float(n * (n - 1))
    sum_xy = np.mean(k_xy)

    return float(sum_xx + sum_yy - 2.0 * sum_xy)


def calculate_kid_from_features(fake_features, real_features, subsets=50, subset_size=50):
    rng = np.random.RandomState(RANDOM_SEED)

    n_fake = fake_features.shape[0]
    n_real = real_features.shape[0]

    subset_size = min(subset_size, n_fake, n_real)

    if subset_size < 2:
        raise RuntimeError("KID requires at least 2 generated and 2 real images.")

    scores = []

    for _ in range(subsets):
        fake_idx = rng.choice(n_fake, subset_size, replace=False)
        real_idx = rng.choice(n_real, subset_size, replace=False)

        x = fake_features[fake_idx]
        y = real_features[real_idx]

        scores.append(mmd2_unbiased(x, y))

    return float(np.mean(scores)), float(np.std(scores))


def compute_custom_fid_kid(fake_paths, real_paths):
    rows = []

    print("\n" + "=" * 80)
    print("Computing custom FID/KID without torchmetrics")
    print("=" * 80)
    print("[info] generated image count: {}".format(len(fake_paths)))
    print("[info] real reference image count: {}".format(len(real_paths)))

    if not ENABLE_FID_KID:
        rows.append({
            "metric": "FID/KID",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "disabled",
            "status": "skipped",
            "note": "ENABLE_FID_KID=False"
        })
        return pd.DataFrame(rows)

    if len(fake_paths) < 2 or len(real_paths) < 2:
        rows.append({
            "metric": "FID/KID",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": "Need at least 2 generated and 2 real images."
        })
        return pd.DataFrame(rows)

    extractor = InceptionFeatureExtractor()

    if not extractor.available:
        rows.append({
            "metric": "FID",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": extractor.note
        })
        rows.append({
            "metric": "KID_mean",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": extractor.note
        })
        rows.append({
            "metric": "KID_std",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": extractor.note
        })
        return pd.DataFrame(rows)

    try:
        print("[info] extracting generated-image features")
        fake_features = extractor.extract(fake_paths, batch_size=FID_KID_BATCH_SIZE)

        print("[info] extracting real-image features")
        real_features = extractor.extract(real_paths, batch_size=FID_KID_BATCH_SIZE)

        print("[info] fake feature shape: {}".format(fake_features.shape))
        print("[info] real feature shape: {}".format(real_features.shape))

    except Exception as e:
        note = repr(e)
        rows.append({
            "metric": "FID",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": note
        })
        rows.append({
            "metric": "KID_mean",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": note
        })
        rows.append({
            "metric": "KID_std",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision",
            "status": "failed",
            "note": note
        })
        return pd.DataFrame(rows)

    # FID
    try:
        print("[info] calculating FID")
        fid_value = calculate_fid_from_features(fake_features, real_features)

        rows.append({
            "metric": "FID",
            "value": fid_value,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "success",
            "note": "Folder-level FID. With only 50 generated images, this estimate is noisy."
        })

        print("[result] FID = {:.6f}".format(fid_value))

    except Exception as e:
        rows.append({
            "metric": "FID",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "failed",
            "note": repr(e)
        })
        print("[result] FID failed: {}".format(repr(e)))

    # KID
    try:
        print("[info] calculating KID")

        kid_mean, kid_std = calculate_kid_from_features(
            fake_features,
            real_features,
            subsets=KID_SUBSETS,
            subset_size=KID_SUBSET_SIZE
        )

        rows.append({
            "metric": "KID_mean",
            "value": kid_mean,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "success",
            "note": "Folder-level KID mean. More reliable than FID for small sample sizes."
        })

        rows.append({
            "metric": "KID_std",
            "value": kid_std,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "success",
            "note": "KID standard deviation over random subsets."
        })

        print("[result] KID_mean = {:.6f}".format(kid_mean))
        print("[result] KID_std  = {:.6f}".format(kid_std))

    except Exception as e:
        rows.append({
            "metric": "KID_mean",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "failed",
            "note": repr(e)
        })
        rows.append({
            "metric": "KID_std",
            "value": np.nan,
            "direction": "Lower is better",
            "implementation": "custom_torchvision_inception_v3",
            "status": "failed",
            "note": repr(e)
        })
        print("[result] KID failed: {}".format(repr(e)))

    for r in rows:
        r["fake_folder"] = str(IMAGE_DIR)
        r["real_folder"] = str(REAL_IMAGE_DIR)
        r["num_fake"] = len(fake_paths)
        r["num_real"] = len(real_paths)

    return pd.DataFrame(rows)


# ============================================================
# DataFrame helpers
# ============================================================

def ensure_columns(df, columns):
    for col in columns:
        if col not in df.columns:
            df[col] = np.nan
    return df


def add_average_row(df):
    avg_row = {}

    for col in df.columns:
        if col == "image_name":
            avg_row[col] = "AVERAGE"
        elif col in ["image_path", "prompt"]:
            avg_row[col] = ""
        elif pd.api.types.is_numeric_dtype(df[col]):
            avg_row[col] = df[col].mean(skipna=True)
        else:
            avg_row[col] = ""

    return pd.concat([df, pd.DataFrame([avg_row])], ignore_index=True)


def build_primary_metric_df(df):
    cols = ["image_name", "image_path", "prompt"] + PRIMARY_METRIC_COLUMNS
    cols = [c for c in cols if c in df.columns]
    return df[cols].copy()


def build_auxiliary_metric_df(df):
    base_cols = ["image_name", "image_path", "prompt"]

    aux_cols = []
    for c in df.columns:
        if c in base_cols:
            continue
        if c in PRIMARY_METRIC_COLUMNS:
            continue
        if c.startswith("pyiqa_"):
            continue
        aux_cols.append(c)

    cols = base_cols + aux_cols
    cols = [c for c in cols if c in df.columns]
    return df[cols].copy()


def build_metric_notes_df():
    rows = [
        ["NIQE", "pyiqa_niqe", "Primary NR-IQA", "Lower is better", "No", "No", "Natural-statistics no-reference metric."],
        ["BRISQUE", "pyiqa_brisque", "Primary NR-IQA", "Lower is better", "No", "No", "Blind/referenceless spatial quality metric."],
        ["PI", "pyiqa_pi", "Primary perceptual index", "Lower is better", "No", "No", "Often used for restoration/generation."],
        ["NRQM", "pyiqa_nrqm", "Primary NR-IQA", "Higher is usually better", "No", "No", "No-reference quality metric."],
        ["MUSIQ", "pyiqa_musiq", "Learned NR-IQA", "Higher is usually better", "No", "No", "Transformer-based perceptual quality metric."],
        ["MANIQA", "pyiqa_maniqa", "Learned NR-IQA", "Higher is usually better", "No", "No", "Learned image quality metric."],
        ["CLIP-IQA", "pyiqa_clipiqa / pyiqa_clipiqa+", "CLIP-based NR-IQA", "Higher is usually better", "No", "No", "CLIP-based image quality score."],
        ["DB-CNN", "pyiqa_dbcnn", "Learned NR-IQA", "Higher is usually better", "No", "No", "Deep bilinear CNN quality score."],
        ["NIMA", "pyiqa_nima", "Aesthetic/perceptual metric", "Higher is usually better", "No", "No", "Neural image assessment."],
        ["HyperIQA", "pyiqa_hyperiqa", "Learned NR-IQA", "Higher is usually better", "No", "No", "Learned no-reference quality score."],
        ["TOPIQ-NR", "pyiqa_topiq_nr", "Learned NR-IQA", "Higher is usually better", "No", "No", "Modern no-reference IQA metric."],
        ["LIQE", "pyiqa_liqe", "Language-image quality evaluator", "Higher is usually better", "No", "No", "Semantic/perceptual no-reference metric."],
        ["CLIPScore", "clipscore", "Prompt-image alignment", "Higher is better", "Yes", "No", "Uses global prompt."],
        ["ImageReward", "imagereward", "Human preference reward", "Higher is better", "Yes", "No", "Uses global prompt."],
        ["PickScore", "pickscore", "Preference score", "Higher is better", "Yes", "No", "Uses global prompt."],
        ["HPSv2", "hpsv2", "Human preference score", "Higher is better", "Yes", "No", "Uses global prompt."],
        ["FID", "distribution_metrics", "Folder-level distribution metric", "Lower is better", "No", "Yes", "Not per-image. Needs real reference folder."],
        ["KID", "distribution_metrics", "Folder-level distribution metric", "Lower is better", "No", "Yes", "Not per-image. Needs real reference folder."],
        ["Auxiliary diagnostics", "auxiliary_diagnostics", "Low-level statistics", "Depends", "No", "No", "Use as supplementary/internal diagnostics."],
    ]

    return pd.DataFrame(
        rows,
        columns=[
            "metric",
            "column",
            "type",
            "direction",
            "requires_prompt",
            "requires_real_reference_folder",
            "note"
        ]
    )


def print_metric_summary(df):
    print("\n" + "=" * 80)
    print("Per-image primary metric averages")
    print("=" * 80)

    for col in PRIMARY_METRIC_COLUMNS:
        if col not in df.columns:
            print("{:<24s} {}".format(col, "missing"))
            continue

        values = pd.to_numeric(df[col], errors="coerce")
        valid_count = values.notna().sum()

        if valid_count == 0:
            print("{:<24s} {}".format(col, "NaN / unavailable"))
        else:
            print("{:<24s} mean={:.6f}   valid={}/{}".format(
                col,
                float(values.mean(skipna=True)),
                int(valid_count),
                len(values)
            ))

    print("=" * 80)


# ============================================================
# Excel output
# ============================================================

def write_sheet(writer, df, sheet_name):
    df.to_excel(writer, index=False, sheet_name=sheet_name)

    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        avg_fill = PatternFill("solid", fgColor="FFF2CC")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if worksheet.max_row >= 2:
            first_value = worksheet.cell(row=worksheet.max_row, column=1).value
            if str(first_value).strip().upper() == "AVERAGE":
                for cell in worksheet[worksheet.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = avg_fill

    except Exception as e:
        print("[warning] Excel styling failed for {}: {}".format(sheet_name, repr(e)))

    try:
        from openpyxl.utils import get_column_letter

        for col_idx, col_cells in enumerate(worksheet.columns, start=1):
            max_len = 0

            for cell in col_cells:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = min(max_len + 2, 70)

    except Exception as e:
        print("[warning] Auto width failed for {}: {}".format(sheet_name, repr(e)))


def save_excel(df_all, distribution_df, metric_status_df, output_path):
    primary_df = build_primary_metric_df(df_all)
    auxiliary_df = build_auxiliary_metric_df(df_all)
    notes_df = build_metric_notes_df()

    df_all_avg = add_average_row(df_all)
    primary_df_avg = add_average_row(primary_df)
    auxiliary_df_avg = add_average_row(auxiliary_df)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        write_sheet(writer, df_all_avg, "all_metrics")
        write_sheet(writer, primary_df_avg, "primary_paper_metrics")
        write_sheet(writer, auxiliary_df_avg, "auxiliary_diagnostics")
        write_sheet(writer, distribution_df, "distribution_metrics")
        write_sheet(writer, metric_status_df, "metric_status")
        write_sheet(writer, notes_df, "metric_notes")


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 80)
    print("Full perceptual metric computation")
    print("=" * 80)
    print("[info] generated folder: {}".format(IMAGE_DIR))
    print("[info] real folder: {}".format(REAL_IMAGE_DIR))
    print("[info] output Excel: {}".format(OUTPUT_XLSX))
    print("[info] global prompt: {}".format(GLOBAL_PROMPT))

    image_paths = collect_images(IMAGE_DIR, IMAGE_EXTS, recursive=False)
    real_paths = collect_images(REAL_IMAGE_DIR, REAL_IMAGE_EXTS, recursive=RECURSIVE_REAL_SCAN)

    print("[info] generated image count: {}".format(len(image_paths)))
    print("[info] real reference image count: {}".format(len(real_paths)))

    if len(image_paths) == 0:
        raise RuntimeError("No generated images found.")

    metric_status_rows = []

    pyiqa_metrics, pyiqa_status = init_pyiqa_metrics(PYIQA_NR_METRICS)
    metric_status_rows.extend(pyiqa_status)

    prompt_scorers, prompt_status = init_prompt_scorers()
    metric_status_rows.extend(prompt_status)

    rows = []

    for idx, image_path in enumerate(image_paths, start=1):
        print("\n[{}/{}] processing: {}".format(idx, len(image_paths), image_path.name))

        row = {
            "image_name": image_path.name,
            "image_path": str(image_path),
            "prompt": GLOBAL_PROMPT,
        }

        try:
            row.update(compute_auxiliary_metrics(image_path))
        except Exception as e:
            print("[error] auxiliary metrics failed on {}: {}".format(image_path.name, repr(e)))

        row.update(compute_pyiqa_metrics(image_path, pyiqa_metrics))
        row.update(compute_prompt_metrics(image_path, prompt_scorers))

        rows.append(row)

    df_all = pd.DataFrame(rows)

    # Force all expected paper metric columns to exist.
    df_all = ensure_columns(df_all, PRIMARY_METRIC_COLUMNS)

    print_metric_summary(df_all)

    distribution_df = compute_custom_fid_kid(image_paths, real_paths)

    # Add FID/KID status to metric_status sheet.
    for _, r in distribution_df.iterrows():
        metric_status_rows.append({
            "metric": r.get("metric", ""),
            "group": "distribution",
            "status": r.get("status", ""),
            "note": r.get("note", "")
        })

    metric_status_df = pd.DataFrame(metric_status_rows)

    save_excel(df_all, distribution_df, metric_status_df, OUTPUT_XLSX)

    print("\n" + "=" * 80)
    print("[done] Excel saved:")
    print(OUTPUT_XLSX)
    print("=" * 80)

    print("\n[Excel sheets]")
    print("  all_metrics              : every per-image metric and diagnostic")
    print("  primary_paper_metrics    : main per-image paper metrics")
    print("  auxiliary_diagnostics    : low-level statistics")
    print("  distribution_metrics     : FID/KID")
    print("  metric_status            : which metrics loaded/failed/skipped")
    print("  metric_notes             : metric meanings and directions")


if __name__ == "__main__":
    main()
